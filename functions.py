import pandas as pd
import os
import pyodbc  # For database connection
from openpyxl import load_workbook
from openpyxl.styles import Font


#too many parameters, a simpler dictionary would be better.
def execute_query_with_excel_data(excel_file_path, column_name, query_name, db_connection_string, output_file_path, in_clause_column):
    try:
        # Step 1: Read Excel and get the custom column
        df = pd.read_excel(excel_file_path)
        if column_name not in df.columns:
            raise ValueError(f"Column '{column_name}' not found in the Excel sheet.")
        
        column_values = df[column_name].dropna().astype(str).tolist()
        if not column_values:
            raise ValueError(f"The column '{column_name}' is empty.")
        
        # Step 2: Load the SQL query
        query_file_path = os.path.join("./queryfolder", query_name)
        if not os.path.exists(query_file_path):
            raise FileNotFoundError(f"Query file '{query_name}' not found in './queryfolder'.")
        
        with open(query_file_path, "r") as query_file:
            query_template = query_file.read()
        
        # Step 3: Prepare the `IN` clause
        in_clause = ", ".join(f"'{value}'" for value in column_values)
        query = query_template.replace(f"{in_clause_column} IN ()", f"{in_clause_column} IN ({in_clause})")
        
        # Step 4: Connect to the database and execute the query
        conn = pyodbc.connect(db_connection_string)  
        cursor = conn.cursor()
        cursor.execute(query)
        result = cursor.fetchall()
        
        # Step 5: Fetch column names
        columns = [desc[0] for desc in cursor.description]
        
        # Step 6: Export results to Excel with bold headers
        result_df = pd.DataFrame(result, columns=columns)
        result_df.to_excel(output_file_path, index=False, engine="openpyxl")
        
        # Bold the header in the exported Excel file
        wb = load_workbook(output_file_path)
        ws = wb.active
        
        # Apply bold to the header row (first row)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        wb.save(output_file_path)
        print(f"Query executed successfully. Results exported to '{output_file_path}'.")
    
    except Exception as e:
        print(f"An error occurred: {e}")

# Creating Hyper file from output excel to display it in  Tableau
from tableauhyperapi import HyperProcess, Telemetry, Connection, TableDefinition, SqlType, CreateMode, Telemetry, \
    SqlQuery, TableColumn, HyperException, Table

def export_to_tableau_hyper(output_file_path, hyper_output_path):
    try:
        df = pd.read_excel(output_file_path)
        
        with HyperProcess(telemetry=Telemetry.SEND_USAGE_DATA_TO_TABLEAU) as hyper:
            with Connection(endpoint=hyper.endpoint, create_mode=CreateMode.CREATE_AND_REPLACE, path=hyper_output_path) as connection:
                
                # Create a table in the Hyper file 
                table_definition = TableDefinition(table_name="Extract", columns=[
                    TableColumn(column_name=col, sql_type=SqlType.text()) for col in df.columns
                ])
                connection.catalog.create_table(table_definition)
                
                # Insert the data into the Hyper file
                rows_to_insert = df.values.tolist()
                connection.execute_sql(
                    SqlQuery(f"INSERT INTO Extract ({', '.join(df.columns)}) VALUES ({', '.join(['?' for _ in df.columns])})"),
                    rows_to_insert
                )
        print(f"Data successfully exported to Tableau Hyper file: {hyper_output_path}")
    
    except HyperException as e:
        print(f"Error creating Hyper file: {e}")

# USE IT HERE add the values
execute_query_with_excel_data(
    excel_file_path="",
    column_name="",
    query_name="",  
    db_connection_string="",
    output_file_path="",
    in_clause_column=""
)

# to create  .hyper from excel if that is needed 
export_to_tableau_hyper(output_file_path="", hyper_output_path="")
